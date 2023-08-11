
import pdb
import numpy as np
from datetime import datetime
import json
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# Reporte PDF
from django.http import HttpResponse
import io
import matplotlib
import mplstereonet
import numpy as np
import matplotlib.pyplot as plt
from shapely.geometry import Point, Polygon, LineString
from fpdf import FPDF
from PDFconFPDF import create_pdf
def clean_data_frame(data, n_decimal):

    data['cota'] = pd.to_numeric(data['cota'], errors='coerce')
    data = data.replace(-1, '-')
    if 'DIPDIR' in data.columns:
        data['DIPDIR'] = pd.to_numeric(data['DIPDIR'], errors='coerce')
    data = data.replace('', '-')
    data = data.round(n_decimal).fillna('-')    
    data.index = np.arange(1, len(data)+1)

    return data
def create_report_FPDF( xlsx_File = "plano-niv.xlsx", mina = "ejemplo" , user  = "Hibring" , rangoCotas=(0, 0), reportType = 'plano', fileType= 'report',  nombrePlano   = "GeoRec-Nivel-7", opTerzaghi= '1'):

    #Leyendo archivo xlsx con las diaclasas y las fallas
   
    file = ''
    mina = mina.replace('/', '-')

    if reportType == 'plano':
        file = nombrePlano
    else:
        file = mina

    data_xlsx = openpyxl.load_workbook(xlsx_File, read_only=True)

    diaclasas = pd.read_excel(xlsx_File , sheet_name='Diaclasas')
    fallas =  pd.read_excel(xlsx_File , sheet_name='Fallas')
    galerias =  pd.read_excel(xlsx_File , sheet_name='Galerias')
    weights = pd.read_excel(xlsx_File , sheet_name='metadata')

    diaclasas = clean_data_frame(diaclasas, 4)
    fallas = clean_data_frame(fallas, 4)

    #eliminando diaclasas subhorizontales- porque estas no se gráfican
    diaclasas =  diaclasas [ diaclasas ['x2'] != '-']
    diaclasas.reset_index(drop=True, inplace=True)
    fallas.reset_index(drop=True, inplace=True)

    #asignando pesos que están en metadata
    diaclasas['weight']= weights['Diaclasas']
    fallas['weight']= weights['Fallas']
    
    
    diaclasas['cota'] = pd.to_numeric(diaclasas['cota'].replace('-', '-1'))
    diaclasas = diaclasas.sort_values(by=['cota', 'weight'])

    fallas['cota'] = pd.to_numeric(fallas['cota'].replace('-', '-1'))
    fallas = fallas.sort_values(by=['cota', 'weight'])

    if rangoCotas != (0, 0):

        rangoInf = rangoCotas[0]
        rangoSup = rangoCotas[1]
        diaclasas = diaclasas.query(
            "`cota` >= @rangoInf and `cota` <= @rangoSup ")
        fallas = fallas.query("`cota` >= @rangoInf and `cota` <= @rangoSup ")

        if diaclasas.empty and fallas.empty:
            print("elementos no encontrados")

            print('X-Error-sin_elementos')
            return 

    puntos = []

    if diaclasas.empty and fallas.empty:
        print("elementos no encontrados")
        mime_type = "('application/pdf', None)"
        return HttpResponse(content_type=mime_type, headers={'X-Error': 'sin_elementos'})

    x = pd.concat([diaclasas['x1'], diaclasas['x3'],
                  fallas['x1'], fallas['x3']])
    y = pd.concat([diaclasas['y1'], diaclasas['y3'],
                  fallas['y1'], fallas['y3']])

    # Seleccionando el recuadro en donde se están todos los objetos encontrados
    coord_x_min = x.min()
    coord_y_max = y.max()
    coord_x_max = x.max()  # round((coord_x1 + x.max())/2,2)
    coord_y_min = y.min()

    # Normalizar weights
    diaclasas['weight'] = pd.to_numeric(diaclasas['weight'].replace('-', '0'))
    diaclasas['weight'] = diaclasas['weight'].replace(
        0, diaclasas[diaclasas['weight'] > 0]['weight'].mean())

    fallas['weight'] = pd.to_numeric(fallas['weight'].replace('-', '0'))
    fallas['weight'] = fallas['weight'].replace(
        0, fallas[fallas['weight'] > 0]['weight'].mean())

    cotasTable = pd.concat([fallas['cota'], diaclasas['cota']])
    cotasTable.replace(-1, np.nan, inplace=True)
    
    d_coord_max = {'x': str(round(coord_x_max, 2)),
                    'y': str(round(coord_y_max, 2))}
    
    d_coord_min = {'x': str(round(coord_x_min,2)),
                    'y': str(round(coord_y_min,2))}
    cotas=[cotasTable.min(), cotasTable.max()]
    if fileType == 'XLS':
        pass
        # return create_report_XLS(diaclasas, fallas, galerias, file)
    
    #aqui hacer el fpdf
    elif fileType == 'report':
        pdf = create_pdf()
        pdf.informe(diaclasas, fallas, galerias, file, mina, user, rangoCotas, reportType, nombrePlano, opTerzaghi,cotas, d_coord_max, d_coord_min)
        pdf.output("informe.pdf")
    
    return 0


def get_data(totalRows, data, opTerzaghi):
    try:
        errorAzimut = data['DIPDIR'].value_counts(dropna=False)['-']
    except:
        errorAzimut = 0
    try:
        errorFrecuencia = data['DIP'].value_counts(dropna=False)['-']

    except:
        errorFrecuencia = 0

    try:
        errorCotas = data['cota'].value_counts(dropna=False)[-1]
    except:
        errorCotas = 0
    eficiencia = (
        (3*totalRows - (errorAzimut + errorFrecuencia+errorCotas)) / (3*totalRows))*100
    eficiencia = round(eficiencia, 2)
    percCotas = round((((totalRows - errorCotas) / totalRows)*100), 2)
    percAzimut = round((((totalRows - errorAzimut) / totalRows)*100), 2)
    percFrecuencia = round(
        (((totalRows - errorFrecuencia) / totalRows)*100), 2)

    data = data[data['DIP'] != "-"]
    data.loc[:, 'DIP'] = data['DIP'].astype('float', errors='ignore')
    data.loc[:, 'DIPDIR'] = data['DIPDIR'].astype('float', errors='ignore')

    # data['DIP'] = data['DIP'].astype('float', errors='ignore')
    # data['DIPDIR'] = data['DIPDIR'].astype('float', errors='ignore')

    #transformación de strikes para que queden entre 0 y
    strikes = np.array([dd-90 if dd-90 >= 0 else dd +
                    270 for dd in data['DIPDIR']])
    # strikes= np.array([dd+90 if dd+90<=360 else dd-270 for dd in data['DIPDIR']])

    modaDipdir, modaDip, bufHistogramas = make_histogram(strikes, data['DIPDIR'].astype(
        int).values, data['DIP'].astype(int).values, data['weight'], opTerzaghi)
    bufHistogramas.seek(0)
    # graficos de polo y diagrama de rosa
    modaStrikes, bufDiagrams = make_RoseDiagrams(
        strikes, data['DIP'].astype(int).values, data['weight'], opTerzaghi)
    bufDiagrams.seek(0)

    maxDipdir = get_intervalo(data['DIPDIR'].max(skipna=True), 0, 360, 20)
    minDipdir = get_intervalo(data['DIPDIR'].min(skipna=True), 0, 360, 20)
    maxDip = get_intervalo(data['DIP'].max(skipna=True), 0, 90, 10)
    minDip = get_intervalo(data['DIP'].min(skipna=True), 0, 90, 10)
    
    return eficiencia, percAzimut, percFrecuencia, percCotas, modaDipdir, modaDip, bufHistogramas, modaStrikes, bufDiagrams, maxDipdir, minDipdir, maxDip, minDip

class create_pdf(FPDF):
    
    def informe(self,diaclasas, fallas, galerias, file, mina, user, rangoCotas, reportType, nombrePlano, opTerzaghi,cotas,cotasmin,cotasmax):
        
        self.set_title("Informe GeoRec")
        #define ancho de las lineas y color negro
        self.set_draw_color(0,0,0) 
        self.set_line_width(0.4)
        
        #agrega fuentes necesarias para el pdf
        self.add_fonts()
        
        #titulo
        self.pdf_title("Informe Levantamiento Estructural")
        
        #tabla datos del informe
        self.metadata(user, opTerzaghi, cotas, cotasmin, cotasmax, nombrePlano)
        
        #tabla diaclasas
        self.data_table(diaclasas, "Diaclasas",opTerzaghi)
        
        #tabla fallas
        self.data_table(fallas, "Fallas",opTerzaghi)
        
    def add_fonts(self):
        self.add_font('NexaRegular', '', "fonts/Nexa Regular.ttf", uni=True)
        self.add_font('NexaBold', '', "fonts/Nexa Bold.ttf", uni=True)
        self.add_font('NexaLight', '', "fonts/Nexa Light.otf", uni=True)
        self.add_font('FuturaMedium', '', "fonts/Futura book font.ttf", uni=True)
        self.add_font('FuturaBold', '', "fonts/Futura medium bold.ttf", uni=True)
    
    def pdf_title(self, title):
        self.add_page()

        self.set_font('NexaBold', '', 30)
        self.set_text_color(0, 0, 0)
        self.set_x(10)
        self.cell(self.w, 20, title, align="L")
        self.ln(2)
     
    def metadata(self, user, opTerzaghi, cotas, cotasmin, cotasmax,nombrePlano):
        #separacion titulo
        self.ln(17)
        metadata_width = self.w - 20

        #guarda la posicion para el titulo del plano
        self.set_font('FuturaBold', '', 12)
        current_x = self.get_x()
        current_y = self.get_y()
        
        #tabla con  metadata
        with self.table(width= metadata_width, col_widths=(1.5, 0.5, 1, 1), line_height=10, first_row_as_headings=False) as table:
            row = table.row()
            row.cell("Plano: ",colspan=2)
            self.set_font('FuturaMedium', '', 12)
            row.cell("void")
            self.set_font('FuturaMedium', '', 8)
            row.cell(f" ( {cotasmax['x']} , {cotasmin['y']} )")
            row.cell(f" ( {cotasmin['x']} , {cotasmin['y']} )")
        
        #se inserta el titulo del plano
        # Se hace así para que esté en formatos distintos (negrita y no-negrita)
        aux_x = self.get_x()
        aux_y = self.get_y()
        self.set_xy(current_x, current_y)
        self.set_font('FuturaMedium', '', 12)
        self.cell(self.get_string_width("Plano: ")+self.get_string_width(nombrePlano)+3, 10, nombrePlano, align="R", border=0)
        self.set_xy(aux_x, aux_y) #vuelvo a la posicion original del puntero (donde debe ir el resto del informe)

        #estilo
        self.set_font('FuturaMedium', '', 10)
        #decido si opTergazhi o no
        if(opTerzaghi):
            estado="Activado"
        else:
            estado="Desactivado"
        
        #tabla con  metadata
        with self.table(width= metadata_width, col_widths=(1.5, 0.5, 1, 1), line_height=10, first_row_as_headings=False, text_align="CENTER" ) as table:
            row = table.row()
            current_x = self.get_x() #se guarda donde debe ir el titulo de cotas
            current_y = self.get_y()
            row.cell(f"\nMin: {cotas[0]}  Máx: {cotas[1]}\n")
            t_current_x = self.get_x() + metadata_width*1.5/4 #y el titulo de Terzaghi
            t_current_y = self.get_y()
            row.cell(f"\n{estado}\n")
        
        #se inserta titulo de cotas    
        aux_y = self.get_y()
        self.set_xy(current_x, current_y)
        self.set_font('FuturaBold', '', 12)
        self.cell(metadata_width*1.5/4, 10, "Cotas", align="C", border=0)
        
        #lo mismo con Terzaghi
        self.set_xy(t_current_x, t_current_y)
        self.cell(metadata_width*0.5/4, 10, "Terzaghi", align="C", border=0)
        
        #se regresa al punto donde se estaba en el informe
        self.set_y(aux_y - 20)
        self.set_x(metadata_width/2 + 10)
        
        #estilos
        self.set_font('FuturaMedium', '', 8)
        
        #resto de datos
        with self.table(width=metadata_width/2,col_widths=(1,1),line_height=10, align="LEFT" , first_row_as_headings=False ) as table:
            row = table.row()
            row.cell(f" ( {cotasmin['x']} , {cotasmax['y']} )")
            row.cell(f" ( {cotasmax['x']} , {cotasmax['y']} )")
            
            row = table.row()
            from datetime import date
            fecha_hoy = date.today().strftime("%d/%m/%Y")
            row.cell(f"Fecha: {fecha_hoy}")
            row.cell(f"Usuario: {user}")
     
    def data_table(self, data, title, opTerzaghi):
        #titulo
        totalRows = len(data.index)
        self.set_font('NexaRegular', '', 15)
        self.cell(self.w - 20, 20, f"{title} ({totalRows} en total)", align="L")
        self.ln(5)
        
        #conseguir datos y graficos
        (eficiencia,
         percAzimut,
         percFrecuencia,
         percCotas,
         modaDipdir,
         modaDip,
         bufHistogramas,
         modaStrikes,
         bufDiagrams,
         maxDipdir,
         minDipdir,
         maxDip,
         minDip) = get_data(totalRows, data, opTerzaghi)
        
        width_aux = (self.w - 20)/5
        width_left = width_aux*2
        width_right = width_aux*3
        
        #estilos tabla
        self.set_font('Times', '', 10)
        self.ln(10)
        #guardar posicion altura
        current_x = self.get_x()
        current_y = self.get_y()
        self.set_font("FuturaMedium", '',11)
        
        #tabla con eficiencia, dipdir, frecuencia, cotas
        with self.table(width=width_left,col_widths=(1,1),line_height=12, align="LEFT", first_row_as_headings=False) as table:
            row = table.row()
            row.cell(f"Eficiencia: {eficiencia}%")
            row.cell(f"Dipdir: {percAzimut}%")
            
            row = table.row()
            row.cell(f"Frecuencia: {percFrecuencia}%")
            row.cell(f"Cotas: {percCotas}%")
            
        
        self.set_y(current_y)
        self.set_x(width_left + 10)
        
        #tabla con datos dip
        with self.table(width=width_right,col_widths=(1,1,1), line_height=6, align="LEFT", text_align="LEFT", first_row_as_headings=False) as table:
            row = table.row()
            row.cell(f"Dip\nMínimo: [{minDip[0]},{minDip[1]}]\nMáximo: [{maxDip[0]},{maxDip[1]}]\nModa: [{modaDip[0]},{modaDip[1]}]")
            row.cell(f"Dip\nMínimo: [{minDip[0]},{minDip[1]}]\nMáximo: [{maxDip[0]},{maxDip[1]}]\nModa: [{modaDip[0]},{modaDip[1]}]")
            row.cell(f"Dip\nMínimo: [{minDip[0]},{minDip[1]}]\nMáximo: [{maxDip[0]},{maxDip[1]}]\nModa: [{modaDip[0]},{modaDip[1]}]")
        
        #Logica para insercion de imagenes y titulos
        current_y = self.get_y()
        current_x = self.get_x()
        self.image(bufHistogramas, w=width_left-5, h=width_left-15)
        self.set_xy(current_x,current_y)
        self.cell(width_left,width_left-15,"",align="C", border=1)
        self.set_xy(width_left + 10, current_y+5)
        current_x = self.get_x()
        self.image(bufDiagrams, w=width_right)
        self.set_xy(current_x,current_y+3)
        self.set_font("FuturaMedium",'',8)
        self.cell(width_right, 0, "Concentración de polos                            Rosetas de rumbo", align="C", border=0)
        self.set_xy(current_x,current_y)
        self.cell(width_right,width_left-15,"",align="C", border=1)
        self.ln(width_left-20)
            
    def header(self):
        self.set_draw_color(0,0,0)
        self.set_line_width(0.4)
        self.set_font('FuturaBold', '', 10)
        self.set_text_color(128, 128, 128)
        self.set_xy(15,10)
        self.cell(0, 10, 'GEOREC | ejemplo', 0, 0, 'L')
        self.set_xy(self.w - 15 - 12, 5)
        # now the logo to the right
        self.image('img/Logo.png', w=12, h=12)
        self.ln(1.5)
        self.set_x(15)
        # now the line
        self.cell(self.w - 30, 0, '', 'T', 2)
    
    
    
    